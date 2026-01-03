"""  
Bug Reporting Tool - Flask Application
A professional bug tracking system with authentication and full CRUD operations

Author: Professional Python Developer
Version: 3.0.0
Date: January 1, 2026
"""

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_from_directory
import sqlite3
import hashlib
import os
import secrets
import logging
from datetime import datetime, timedelta
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import re
from contextlib import contextmanager
import base64
import io
from google import genai
from PIL import Image
import requests
import json
from requests_oauthlib import OAuth2Session
from urllib.parse import urlencode

# Initialize Flask app
app = Flask(__name__)

# Configuration with environment variables (production-ready)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV') == 'production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24 hours

# Database configuration
DATABASE = os.environ.get('DATABASE_PATH', 'bug_tracker.db')

# Upload configuration
UPLOAD_FOLDER = os.path.join('static', 'uploads')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Ensure upload folder exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# OAuth Configuration
GOOGLE_OAUTH_CONFIG = {
    'client_id': os.environ.get('GOOGLE_CLIENT_ID', ''),
    'client_secret': os.environ.get('GOOGLE_CLIENT_SECRET', ''),
    'redirect_uri': os.environ.get('GOOGLE_REDIRECT_URI', 'http://localhost:5000/auth/google/callback'),
    'auth_uri': 'https://accounts.google.com/o/oauth2/v2/auth',
    'token_uri': 'https://oauth2.googleapis.com/token',
    'scopes': ['openid', 'email', 'profile']
}

GITHUB_OAUTH_CONFIG = {
    'client_id': os.environ.get('GITHUB_CLIENT_ID', ''),
    'client_secret': os.environ.get('GITHUB_CLIENT_SECRET', ''),
    'redirect_uri': os.environ.get('GITHUB_REDIRECT_URI', 'http://localhost:5000/auth/github/callback'),
    'auth_uri': 'https://github.com/login/oauth/authorize',
    'token_uri': 'https://github.com/login/oauth/access_token',
    'scopes': ['user:email']
}

# Logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bug_tracker.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

@contextmanager
def get_db_connection():
    """Context manager for database connections (best practice)"""
    conn = None
    try:
        conn = sqlite3.connect(DATABASE, timeout=10.0)
        conn.row_factory = sqlite3.Row
        conn.execute('PRAGMA foreign_keys = ON')  # Enable foreign key support
        yield conn
        conn.commit()
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Database error: {str(e)}")
        raise
    finally:
        if conn:
            conn.close()

def init_db():
    """Initialize the database with required tables and indexes"""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Create users table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'debugger',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Create bugs table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS bugs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                description TEXT NOT NULL,
                steps TEXT,
                expected_result TEXT,
                actual_result TEXT,
                screenshot_url TEXT,
                screenshot_path TEXT,
                priority TEXT NOT NULL,
                status TEXT DEFAULT 'Open',
                assigned_to INTEGER,
                created_by INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (created_by) REFERENCES users (id),
                FOREIGN KEY (assigned_to) REFERENCES users (id)
            )
        ''')
        
        # Add expected_result and actual_result columns if they don't exist (migration)
        try:
            cursor.execute('ALTER TABLE bugs ADD COLUMN expected_result TEXT')
            logger.info('Added expected_result column to bugs table')
        except sqlite3.OperationalError:
            pass
        
        try:
            cursor.execute('ALTER TABLE bugs ADD COLUMN actual_result TEXT')
            logger.info('Added actual_result column to bugs table')
        except sqlite3.OperationalError:
            pass
        
        # Add profile fields to users table (migration)
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN full_name TEXT')
            logger.info('Added full_name column to users table')
        except sqlite3.OperationalError:
            pass
        
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN username TEXT')
            logger.info('Added username column to users table')
        except sqlite3.OperationalError:
            pass
        
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN contact_number TEXT')
            logger.info('Added contact_number column to users table')
        except sqlite3.OperationalError:
            pass
        
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN profile_picture TEXT')
            logger.info('Added profile_picture column to users table')
        except sqlite3.OperationalError:
            pass
        
        # Add Gemini avatar generation fields
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN gemini_api_key TEXT')
            logger.info('Added gemini_api_key column to users table')
        except sqlite3.OperationalError:
            pass
        
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN hoodie_color TEXT DEFAULT "#1677ff"')
            logger.info('Added hoodie_color column to users table')
        except sqlite3.OperationalError:
            pass
        
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN gender TEXT')
            logger.info('Added gender column to users table')
        except sqlite3.OperationalError:
            pass
        
        # Add OAuth fields
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN oauth_provider TEXT')
            logger.info('Added oauth_provider column to users table')
        except sqlite3.OperationalError:
            pass
        
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN oauth_id TEXT UNIQUE')
            logger.info('Added oauth_id column to users table')
        except sqlite3.OperationalError:
            pass
        
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN oauth_token TEXT')
            logger.info('Added oauth_token column to users table')
        except sqlite3.OperationalError:
            pass
        
        # Create comments table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS comments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bug_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                comment TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (bug_id) REFERENCES bugs (id),
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        # Create history table for audit trail
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS bug_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bug_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                action TEXT NOT NULL,
                old_value TEXT,
                new_value TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (bug_id) REFERENCES bugs (id),
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        # Create indexes for better performance
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_bugs_status ON bugs(status)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_bugs_priority ON bugs(priority)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_bugs_created_by ON bugs(created_by)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_bugs_assigned_to ON bugs(assigned_to)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_comments_bug_id ON comments(bug_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_users_email ON users(email)')
    
    logger.info("Database initialized successfully with all tables and indexes!")
    print("[OK] Database initialized successfully!")

def hash_password(password):
    """Hash password using werkzeug's secure pbkdf2:sha256"""
    return generate_password_hash(password, method='pbkdf2:sha256', salt_length=16)

def verify_password(stored_hash, password):
    """Verify password against stored hash"""
    return check_password_hash(stored_hash, password)

def validate_email(email):
    """Validate email format"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def format_datetime(dt_string):
    """Format datetime string to human-readable format"""
    try:
        dt = datetime.strptime(dt_string, '%Y-%m-%d %H:%M:%S')
        now = datetime.now()
        diff = now - dt
        
        if diff.days == 0:
            if diff.seconds < 60:
                return "Just now"
            elif diff.seconds < 3600:
                minutes = diff.seconds // 60
                return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
            else:
                hours = diff.seconds // 3600
                return f"{hours} hour{'s' if hours != 1 else ''} ago"
        elif diff.days == 1:
            return "Yesterday"
        elif diff.days < 7:
            return f"{diff.days} days ago"
        else:
            return dt.strftime('%b %d, %Y')
    except:
        return dt_string[:16]

def log_bug_history(conn, bug_id, user_id, action, old_value=None, new_value=None):
    """Log bug history for audit trail"""
    try:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO bug_history (bug_id, user_id, action, old_value, new_value)
            VALUES (?, ?, ?, ?, ?)
        ''', (bug_id, user_id, action, old_value, new_value))
        conn.commit()
        logger.info(f"History logged: {action} for bug #{bug_id}")
    except Exception as e:
        logger.error(f"Error logging history: {str(e)}")

# Register format_datetime as template filter
app.jinja_env.filters['format_datetime'] = format_datetime

def sanitize_input(text, max_length=None):
    """Sanitize user input"""
    if not text:
        return text
    text = text.strip()
    if max_length:
        text = text[:max_length]
    return text

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def admin_required(f):
    """Decorator to require admin role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login to access this page', 'error')
            return redirect(url_for('login'))
        if session.get('user_role') != 'admin':
            flash('Admin access required', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def login_required(f):
    """Decorator to require login for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            logger.warning(f"Unauthorized access attempt to {request.path}")
            flash('Please login to access this page', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ============== AUTHENTICATION ROUTES ==============

@app.route('/')
def index():
    """Home page - redirect based on login status"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    """User registration with comprehensive validation"""
    if request.method == 'POST':
        email = sanitize_input(request.form.get('email', ''), 255)
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        role = request.form.get('role', 'debugger')
        
        # Validate role
        if role not in ['admin', 'debugger']:
            flash('Invalid role selected', 'error')
            return render_template('signup.html')
        
        # Comprehensive validation
        if not email or not password:
            flash('Email and password are required', 'error')
            return render_template('signup.html')
        
        if not validate_email(email):
            flash('Invalid email format', 'error')
            return render_template('signup.html')
        
        if password != confirm_password:
            flash('Passwords do not match', 'error')
            return render_template('signup.html')
        
        if len(password) < 8:
            flash('Password must be at least 8 characters long', 'error')
            return render_template('signup.html')
        
        # Password strength check
        if not any(c.isupper() for c in password) or not any(c.isdigit() for c in password):
            flash('Password must contain at least one uppercase letter and one number', 'error')
            return render_template('signup.html')
        
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                
                # Check if email already exists
                cursor.execute('SELECT id FROM users WHERE email = ?', (email.lower(),))
                if cursor.fetchone():
                    flash('Email already registered. Please login.', 'error')
                    return render_template('signup.html')
                
                # Insert new user with secure password hash and role
                hashed_password = hash_password(password)
                cursor.execute('INSERT INTO users (email, password, role) VALUES (?, ?, ?)',
                             (email.lower(), hashed_password, role))
                
                logger.info(f"New {role} registered: {email}")
                flash(f'Account created successfully as {role.title()}! Please login.', 'success')
                return redirect(url_for('login'))
                
        except Exception as e:
            logger.error(f"Signup error: {str(e)}")
            flash('An error occurred during registration. Please try again.', 'error')
            return render_template('signup.html')
    
    return render_template('signup.html')

@app.route('/auth/google', methods=['GET'])
def google_login():
    """Initiate Google OAuth login"""
    if not GOOGLE_OAUTH_CONFIG['client_id']:
        flash('Google OAuth is not configured. Please contact admin.', 'error')
        return redirect(url_for('login'))
    
    google = OAuth2Session(
        client_id=GOOGLE_OAUTH_CONFIG['client_id'],
        redirect_uri=GOOGLE_OAUTH_CONFIG['redirect_uri'],
        scope=GOOGLE_OAUTH_CONFIG['scopes']
    )
    authorization_url, state = google.authorization_url(
        GOOGLE_OAUTH_CONFIG['auth_uri'],
        access_type='offline',
        prompt='select_account'
    )
    session['oauth_state'] = state
    session['oauth_provider'] = 'google'
    logger.info(f"Google OAuth initiated: {state}")
    return redirect(authorization_url)

@app.route('/auth/google/callback')
def google_callback():
    """Handle Google OAuth callback"""
    try:
        code = request.args.get('code')
        state = request.args.get('state')
        
        if not code or state != session.get('oauth_state'):
            flash('Invalid OAuth state. Please try again.', 'error')
            return redirect(url_for('login'))
        
        # Exchange code for token
        google = OAuth2Session(
            client_id=GOOGLE_OAUTH_CONFIG['client_id'],
            redirect_uri=GOOGLE_OAUTH_CONFIG['redirect_uri'],
            state=state
        )
        token = google.fetch_token(
            GOOGLE_OAUTH_CONFIG['token_uri'],
            client_secret=GOOGLE_OAUTH_CONFIG['client_secret'],
            authorization_response=request.url
        )
        
        # Get user info
        user_info = google.get('https://openidconnect.googleapis.com/v1/userinfo').json()
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Check if user exists
            cursor.execute('SELECT id FROM users WHERE oauth_provider = ? AND oauth_id = ?',
                         ('google', user_info['sub']))
            existing_user = cursor.fetchone()
            
            if existing_user:
                user_id = existing_user['id']
            else:
                # Create new user
                email = user_info.get('email', '').lower()
                full_name = user_info.get('name', '')
                
                # Check if email already exists
                cursor.execute('SELECT id FROM users WHERE email = ?', (email,))
                email_user = cursor.fetchone()
                
                if email_user:
                    user_id = email_user['id']
                else:
                    cursor.execute('''
                        INSERT INTO users (email, password, role, full_name, oauth_provider, oauth_id, oauth_token)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (email, generate_password_hash('oauth'), 'user', full_name, 'google', user_info['sub'], json.dumps(token)))
                    
                    conn.commit()
                    user_id = cursor.lastrowid
            
            # Update OAuth token
            cursor.execute('''
                UPDATE users 
                SET oauth_token = ?, oauth_provider = ?, oauth_id = ?
                WHERE id = ?
            ''', (json.dumps(token), 'google', user_info['sub'], user_id))
            conn.commit()
        
        # Create session
        session['user_id'] = user_id
        session['user_email'] = email
        session['user_role'] = 'user'
        session.permanent = True
        
        logger.info(f"User logged in via Google: {email}")
        flash(f'Welcome, {full_name or email}!', 'success')
        return redirect(url_for('dashboard'))
        
    except Exception as e:
        logger.error(f"Google OAuth error: {str(e)}")
        flash('OAuth login failed. Please try again or use email/password.', 'error')
        return redirect(url_for('login'))

@app.route('/auth/github', methods=['GET'])
def github_login():
    """Initiate GitHub OAuth login"""
    if not GITHUB_OAUTH_CONFIG['client_id']:
        flash('GitHub OAuth is not configured. Please contact admin.', 'error')
        return redirect(url_for('login'))
    
    github = OAuth2Session(
        client_id=GITHUB_OAUTH_CONFIG['client_id'],
        redirect_uri=GITHUB_OAUTH_CONFIG['redirect_uri'],
        scope=GITHUB_OAUTH_CONFIG['scopes']
    )
    authorization_url, state = github.authorization_url(
        GITHUB_OAUTH_CONFIG['auth_uri']
    )
    session['oauth_state'] = state
    session['oauth_provider'] = 'github'
    logger.info(f"GitHub OAuth initiated: {state}")
    return redirect(authorization_url)

@app.route('/auth/github/callback')
def github_callback():
    """Handle GitHub OAuth callback"""
    try:
        code = request.args.get('code')
        state = request.args.get('state')
        
        if not code:
            flash('No authorization code received. Please try again.', 'error')
            return redirect(url_for('login'))
        
        # Exchange code for token
        token_url = GITHUB_OAUTH_CONFIG['token_uri']
        payload = {
            'client_id': GITHUB_OAUTH_CONFIG['client_id'],
            'client_secret': GITHUB_OAUTH_CONFIG['client_secret'],
            'code': code,
            'redirect_uri': GITHUB_OAUTH_CONFIG['redirect_uri']
        }
        headers = {'Accept': 'application/json'}
        
        token_response = requests.post(token_url, json=payload, headers=headers)
        token_data = token_response.json()
        
        if 'error' in token_data:
            flash(f'OAuth error: {token_data.get("error_description", "Unknown error")}', 'error')
            return redirect(url_for('login'))
        
        access_token = token_data.get('access_token')
        
        # Get user info
        user_headers = {'Authorization': f'token {access_token}', 'Accept': 'application/json'}
        user_response = requests.get('https://api.github.com/user', headers=user_headers)
        user_info = user_response.json()
        
        # Get email if not public
        if not user_info.get('email'):
            email_response = requests.get('https://api.github.com/user/emails', headers=user_headers)
            emails = email_response.json()
            user_info['email'] = next((e['email'] for e in emails if e['primary']), emails[0]['email'] if emails else None)
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Check if user exists
            cursor.execute('SELECT id FROM users WHERE oauth_provider = ? AND oauth_id = ?',
                         ('github', str(user_info['id'])))
            existing_user = cursor.fetchone()
            
            if existing_user:
                user_id = existing_user['id']
            else:
                email = user_info.get('email', f"github_{user_info['login']}@github.local").lower()
                full_name = user_info.get('name', user_info.get('login', ''))
                
                # Check if email already exists
                cursor.execute('SELECT id FROM users WHERE email = ?', (email,))
                email_user = cursor.fetchone()
                
                if email_user:
                    user_id = email_user['id']
                else:
                    cursor.execute('''
                        INSERT INTO users (email, password, role, full_name, oauth_provider, oauth_id, oauth_token)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (email, generate_password_hash('oauth'), 'user', full_name, 'github', str(user_info['id']), json.dumps(token_data)))
                    
                    conn.commit()
                    user_id = cursor.lastrowid
            
            # Update OAuth token
            cursor.execute('''
                UPDATE users 
                SET oauth_token = ?, oauth_provider = ?, oauth_id = ?
                WHERE id = ?
            ''', (json.dumps(token_data), 'github', str(user_info['id']), user_id))
            conn.commit()
        
        # Create session
        session['user_id'] = user_id
        session['user_email'] = email
        session['user_role'] = 'user'
        session.permanent = True
        
        logger.info(f"User logged in via GitHub: {email}")
        flash(f'Welcome, {full_name or user_info.get("login", "User")}!', 'success')
        return redirect(url_for('dashboard'))
        
    except Exception as e:
        logger.error(f"GitHub OAuth error: {str(e)}")
        flash('OAuth login failed. Please try again or use email/password.', 'error')
        return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login with secure password verification"""
    if request.method == 'POST':
        email = sanitize_input(request.form.get('email', ''), 255)
        password = request.form.get('password', '')
        
        if not email or not password:
            flash('Email and password are required', 'error')
            return render_template('login.html')
        
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                
                cursor.execute('SELECT id, email, password, role FROM users WHERE email = ?',
                             (email.lower(),))
                user = cursor.fetchone()
                
                if user and verify_password(user['password'], password):
                    session['user_id'] = user['id']
                    session['user_email'] = user['email']
                    session['user_role'] = user['role']
                    session.permanent = True
                    logger.info(f"{user['role'].title()} logged in: {email}")
                    flash(f'Welcome back, {user["role"].title()}!', 'success')
                    return redirect(url_for('dashboard'))
                else:
                    logger.warning(f"Failed login attempt for: {email}")
                    flash('Invalid email or password', 'error')
                    return render_template('login.html')
                    
        except Exception as e:
            logger.error(f"Login error: {str(e)}")
            flash('An error occurred during login. Please try again.', 'error')
            return render_template('login.html')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """User logout"""
    email = session.get('user_email', 'Unknown')
    session.clear()
    logger.info(f"User logged out: {email}")
    flash('Logged out successfully', 'success')
    return redirect(url_for('login'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    """View and edit user profile"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            user_id = session.get('user_id')
            
            if request.method == 'POST':
                full_name = request.form.get('full_name', '').strip()
                username = request.form.get('username', '').strip()
                contact_number = request.form.get('contact_number', '').strip()
                gemini_api_key = request.form.get('gemini_api_key', '').strip()
                hoodie_color = request.form.get('hoodie_color', '#1677ff').strip()
                gender = request.form.get('gender', '').strip()
                
                # Validate contact number format (optional field)
                if contact_number and not re.match(r'^\+?[\d\s\-()]+$', contact_number):
                    flash('Invalid contact number format', 'error')
                    return redirect(url_for('profile'))
                
                # Handle profile picture upload
                profile_picture = None
                if 'profile_picture' in request.files:
                    file = request.files['profile_picture']
                    if file and file.filename and allowed_file(file.filename):
                        filename = secure_filename(f"profile_{user_id}_{secrets.token_hex(8)}_{file.filename}")
                        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                        file.save(filepath)
                        profile_picture = filename
                
                # Update profile
                if profile_picture:
                    cursor.execute('''
                        UPDATE users 
                        SET full_name = ?, username = ?, contact_number = ?, profile_picture = ?,
                            gemini_api_key = ?, hoodie_color = ?, gender = ?
                        WHERE id = ?
                    ''', (full_name, username, contact_number, profile_picture, gemini_api_key, hoodie_color, gender, user_id))
                else:
                    cursor.execute('''
                        UPDATE users 
                        SET full_name = ?, username = ?, contact_number = ?,
                            gemini_api_key = ?, hoodie_color = ?, gender = ?
                        WHERE id = ?
                    ''', (full_name, username, contact_number, gemini_api_key, hoodie_color, gender, user_id))
                
                # Commit the changes to the database
                conn.commit()
                
                # Update session
                session['user_name'] = full_name if full_name else session.get('user_email')
                
                flash('Profile updated successfully!', 'success')
                logger.info(f"Profile updated: User ID {user_id}")
                return redirect(url_for('profile'))
            
            # GET request - fetch current profile data
            cursor.execute('''
                SELECT id, email, role, full_name, username, contact_number, 
                       profile_picture, created_at, gemini_api_key, hoodie_color, gender
                FROM users 
                WHERE id = ?
            ''', (user_id,))
            user = cursor.fetchone()
            
            # Get user statistics
            cursor.execute('SELECT COUNT(*) FROM bugs WHERE created_by = ?', (user_id,))
            bugs_created = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM bugs WHERE assigned_to = ?', (user_id,))
            bugs_assigned = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM comments WHERE user_id = ?', (user_id,))
            comments_count = cursor.fetchone()[0]
            
            return render_template('profile.html', 
                                 user=user, 
                                 bugs_created=bugs_created,
                                 bugs_assigned=bugs_assigned,
                                 comments_count=comments_count)
    
    except Exception as e:
        logger.error(f"Error loading profile: {str(e)}")
        flash('Error loading profile. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/generate-avatar', methods=['POST'])
@login_required
def generate_avatar():
    """Generate animated Ghibli-style avatar using Gemini API"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            user_id = session.get('user_id')
            
            # Get user data
            cursor.execute('''
                SELECT full_name, gemini_api_key, hoodie_color, gender, profile_picture
                FROM users WHERE id = ?
            ''', (user_id,))
            user = cursor.fetchone()
            
            if not user or not user[1]:  # Check if API key exists
                flash('Please add your Gemini API key in your profile first.', 'error')
                return redirect(url_for('profile'))
            
            full_name = user[0] or 'User'
            api_key = user[1]
            hoodie_color = user[2] or '#1677ff'
            gender = user[3] or 'person'
            current_profile_pic = user[4]
            
            # Check if user has a profile picture to use as reference
            has_reference = False
            photo_path = None
            
            if current_profile_pic:
                photo_path = os.path.join(app.config['UPLOAD_FOLDER'], current_profile_pic)
                has_reference = os.path.exists(photo_path)
            
            # Configure Gemini
            client = genai.Client(api_key=api_key)
            
            # Determine text color for name on hoodie (white by default, black for light backgrounds)
            def is_light_color(hex_color):
                """Check if color is light (requires black text)"""
                hex_color = hex_color.lstrip('#')
                r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
                return luminance > 0.7
            
            text_color = "black" if is_light_color(hoodie_color) else "white"
            
            # Create Ghibli-style prompt with photo reference if available
            if has_reference:
                prompt = f"""Create a cute, animated Studio Ghibli-style avatar illustration based on the uploaded photo. 
                
                Style requirements:
                - Studio Ghibli anime art style (like Spirited Away, My Neighbor Totoro)
                - Capture the person's key features (face shape, hair style, distinctive characteristics) from the photo
                - Soft, warm colors with hand-drawn aesthetic
                - Expressive, large eyes with sparkles
                - Gentle smile and friendly expression
                - Wearing a {hoodie_color} colored cozy hoodie/sweatshirt
                - Display the name "{full_name}" on the hoodie in {text_color} text (clean, legible font)
                - Character should be a {gender if gender else 'person'}
                - Simple, clean background (soft gradient or solid pastel color)
                - Wholesome and adorable appearance
                - Head and shoulders portrait orientation
                - Professional quality suitable for profile picture
                
                Art direction: Transform this person into a Ghibli character while maintaining their recognizable features. 
                Capture the magical, whimsical essence of Ghibli characters - innocent, warm, and full of life."""
                
                # Read photo data
                with open(photo_path, 'rb') as f:
                    photo_data = f.read()
                
                # Detect mime type
                mime_type = 'image/jpeg'
                if photo_path.lower().endswith('.png'):
                    mime_type = 'image/png'
                elif photo_path.lower().endswith('.gif'):
                    mime_type = 'image/gif'
                
                # Generate with photo reference
                logger.info(f"Generating avatar for user {user_id} with photo reference")
                
                response = client.models.generate_content(
                    model='gemini-2.0-flash-exp',
                    contents=[
                        prompt,
                        {
                            "inline_data": {
                                "mime_type": mime_type,
                                "data": base64.b64encode(photo_data).decode()
                            }
                        }
                    ]
                )
            else:
                prompt = f"""Create a cute, animated Studio Ghibli-style avatar illustration of {full_name}. 
                
                Style requirements:
                - Studio Ghibli anime art style (like Spirited Away, My Neighbor Totoro)
                - Soft, warm colors with hand-drawn aesthetic
                - Expressive, large eyes with sparkles
                - Gentle smile and friendly expression
                - Wearing a {hoodie_color} colored cozy hoodie/sweatshirt
                - Display the name "{full_name}" on the hoodie in {text_color} text (clean, legible font)
                - Character should be a {gender if gender else 'person'}
                - Simple, clean background (soft gradient or solid pastel color)
                - Wholesome and adorable appearance
                - Head and shoulders portrait orientation
                - Professional quality suitable for profile picture
                
                Art direction: Capture the magical, whimsical essence of Ghibli characters - innocent, warm, and full of life. 
                The character should feel approachable and kind, with that signature Ghibli charm."""
                
                logger.info(f"Generating avatar for user {user_id} without photo reference")
                
                response = client.models.generate_content(
                    model='gemini-2.0-flash-exp',
                    contents=prompt
                )
            
            # Check if image was generated
            if response.text:
                # For text-only responses, we need to use image generation
                # Save the text response or handle accordingly
                flash('Avatar description generated. Image generation requires Imagen model.', 'info')
                logger.info(f"Text response: {response.text[:100]}")
                return redirect(url_for('profile'))
            elif hasattr(response, 'parts') and response.parts:
                for part in response.parts:
                    if hasattr(part, 'inline_data'):
                        # Save the generated image
                        image_data = base64.b64decode(part.inline_data.data) if isinstance(part.inline_data.data, str) else part.inline_data.data
                        filename = secure_filename(f"avatar_{user_id}_{secrets.token_hex(8)}.png")
                        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                        
                        with open(filepath, 'wb') as f:
                            f.write(image_data)
                        
                        # Update user's profile picture
                        cursor.execute('''
                            UPDATE users 
                            SET profile_picture = ?
                            WHERE id = ?
                        ''', (filename, user_id))
                        
                        flash('✨ Avatar generated successfully! Your profile picture has been updated.', 'success')
                        logger.info(f"Avatar generated for user {user_id}: {filename}")
                        return redirect(url_for('profile'))
            
            flash('Could not generate avatar. Please try again or check your API key.', 'error')
            return redirect(url_for('profile'))
            
    except Exception as e:
        logger.error(f"Error generating avatar: {str(e)}")
        flash(f'Error generating avatar: {str(e)}', 'error')
        return redirect(url_for('profile'))

@app.route('/users')
@admin_required
def view_users():
    """View all registered users (Admin only)"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT u.id, u.email, u.role, u.created_at,
                       COUNT(b.id) as bugs_created,
                       COUNT(DISTINCT ba.id) as bugs_assigned
                FROM users u
                LEFT JOIN bugs b ON u.id = b.created_by
                LEFT JOIN bugs ba ON u.id = ba.assigned_to
                GROUP BY u.id
                ORDER BY u.created_at DESC
            ''')
            users = cursor.fetchall()
            
            return render_template('users.html', users=users)
    
    except Exception as e:
        logger.error(f"Error loading users: {str(e)}")
        flash('Error loading users. Please try again.', 'error')
        return redirect(url_for('dashboard'))

# ============== BUG MANAGEMENT ROUTES ==============

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard showing all bugs with statistics"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Get filter parameters
            status_filter = request.args.get('status', '')
            priority_filter = request.args.get('priority', '')
            search_query = sanitize_input(request.args.get('search', ''), 200)
            
            # Build query with filters
            query = '''
                SELECT b.*, 
                       creator.email as creator_email,
                       assignee.email as assignee_email
                FROM bugs b
                LEFT JOIN users creator ON b.created_by = creator.id
                LEFT JOIN users assignee ON b.assigned_to = assignee.id
                WHERE 1=1
            '''
            params = []
            
            if status_filter:
                query += ' AND b.status = ?'
                params.append(status_filter)
            
            if priority_filter:
                query += ' AND b.priority = ?'
                params.append(priority_filter)
            
            if search_query:
                query += ' AND (b.title LIKE ? OR b.description LIKE ?)'
                params.extend([f'%{search_query}%', f'%{search_query}%'])
            
            query += ' ORDER BY b.created_at DESC'
            
            cursor.execute(query, params)
            bugs = cursor.fetchall()
            
            # Get statistics
            cursor.execute('SELECT COUNT(*) as total FROM bugs')
            total_bugs = cursor.fetchone()['total']
            
            cursor.execute("SELECT COUNT(*) as open FROM bugs WHERE status = 'Open'")
            open_bugs = cursor.fetchone()['open']
            
            cursor.execute("SELECT COUNT(*) as in_progress FROM bugs WHERE status = 'In Progress'")
            in_progress_bugs = cursor.fetchone()['in_progress']
            
            cursor.execute("SELECT COUNT(*) as fixed FROM bugs WHERE status = 'Fixed'")
            fixed_bugs = cursor.fetchone()['fixed']
            
            cursor.execute("SELECT COUNT(*) as high FROM bugs WHERE priority = 'High'")
            high_priority = cursor.fetchone()['high']
            
            cursor.execute("SELECT COUNT(*) as closed FROM bugs WHERE status = 'Closed'")
            closed_bugs = cursor.fetchone()['closed']
            
            # Get recent activity (last 10 history records) - with error handling
            recent_activity = []
            notifications = []
            try:
                cursor.execute('''
                    SELECT bh.*, u.email as user_email, b.title as bug_title
                    FROM bug_history bh
                    LEFT JOIN users u ON bh.user_id = u.id
                    LEFT JOIN bugs b ON bh.bug_id = b.id
                    ORDER BY bh.created_at DESC
                    LIMIT 10
                ''')
                recent_activity = cursor.fetchall()

                # Personalized notifications: assignment to me, comments/status on bugs assigned to me
                cursor.execute('''
                    SELECT bh.*, actor.email as actor_email, b.title as bug_title
                    FROM bug_history bh
                    LEFT JOIN users actor ON bh.user_id = actor.id
                    LEFT JOIN bugs b ON bh.bug_id = b.id
                    WHERE (
                        (bh.action = 'assigned_to' AND bh.new_value = ?)
                        OR (b.assigned_to = ? AND bh.action IN ('comment_added', 'status_changed'))
                    )
                    ORDER BY bh.created_at DESC
                    LIMIT 8
                ''', (session.get('user_email'), session.get('user_id')))
                notifications = cursor.fetchall()
            except Exception as activity_error:
                logger.warning(f"Could not load recent activity: {str(activity_error)}")
                recent_activity = []
                notifications = []
            
            # Get all users for assignment dropdown
            cursor.execute('SELECT id, email, role FROM users ORDER BY email')
            users = cursor.fetchall()
            
            # Get unassigned bugs with high priority
            cursor.execute('''
                SELECT b.*, 
                       creator.email as creator_email,
                       assignee.email as assignee_email
                FROM bugs b
                LEFT JOIN users creator ON b.created_by = creator.id
                LEFT JOIN users assignee ON b.assigned_to = assignee.id
                WHERE b.assigned_to IS NULL
                ORDER BY b.priority DESC, b.created_at DESC
                LIMIT 10
            ''')
            unassigned_bugs = cursor.fetchall()
            unassigned_count = len(unassigned_bugs)
            
            # Get high priority bugs
            cursor.execute('''
                SELECT b.*, 
                       creator.email as creator_email,
                       assignee.email as assignee_email
                FROM bugs b
                LEFT JOIN users creator ON b.created_by = creator.id
                LEFT JOIN users assignee ON b.assigned_to = assignee.id
                WHERE b.priority = 'High'
                ORDER BY b.created_at DESC
                LIMIT 10
            ''')
            high_priority_bugs = cursor.fetchall()
            high_priority_count = len(high_priority_bugs)
            
            # Get in-progress bugs
            cursor.execute('''
                SELECT b.*, 
                       creator.email as creator_email,
                       assignee.email as assignee_email
                FROM bugs b
                LEFT JOIN users creator ON b.created_by = creator.id
                LEFT JOIN users assignee ON b.assigned_to = assignee.id
                WHERE b.status = 'In Progress'
                ORDER BY b.created_at DESC
                LIMIT 10
            ''')
            in_progress_bugs_list = cursor.fetchall()
            in_progress_count = len(in_progress_bugs_list)
            
            stats = {
                'total': total_bugs,
                'open': open_bugs,
                'in_progress': in_progress_bugs,
                'fixed': fixed_bugs,
                'high_priority': high_priority,
                'closed': closed_bugs
            }
            
            return render_template('dashboard.html', 
                                 bugs=bugs, 
                                 users=users,
                                 status_filter=status_filter,
                                 priority_filter=priority_filter,
                                 search_query=search_query,
                                 stats=stats,
                                 recent_activity=recent_activity,
                                 notifications=notifications,
                                 unassigned_bugs=unassigned_bugs,
                                 unassigned_count=unassigned_count,
                                 high_priority_bugs=high_priority_bugs,
                                 high_priority_count=high_priority_count,
                                 in_progress_bugs=in_progress_bugs_list,
                                 in_progress_count=in_progress_count)
        
    except Exception as e:
        logger.error(f"Dashboard error: {str(e)}")
        flash('Error loading dashboard. Please try again.', 'error')
        return render_template('dashboard.html', bugs=[], users=[], stats={}, recent_activity=[], notifications=[])

@app.route('/bug/new', methods=['GET', 'POST'])
@login_required
def new_bug():
    """Create a new bug report with validation and image upload"""
    if request.method == 'POST':
        title = sanitize_input(request.form.get('title', ''), 200)
        description = sanitize_input(request.form.get('description', ''), 2000)
        steps = sanitize_input(request.form.get('steps', ''), 2000)
        expected_result = sanitize_input(request.form.get('expected_result', ''), 2000)
        actual_result = sanitize_input(request.form.get('actual_result', ''), 2000)
        screenshot_url = sanitize_input(request.form.get('screenshot_url', ''), 500)
        priority = request.form.get('priority', 'Medium')
        
        # Validation
        if not title or not description:
            flash('Title and description are required', 'error')
            return render_template('new_bug.html')
        
        if len(title) < 5:
            flash('Title must be at least 5 characters long', 'error')
            return render_template('new_bug.html')
        
        if priority not in ['Low', 'Medium', 'High']:
            flash('Invalid priority level', 'error')
            return render_template('new_bug.html')
        
        # Handle file upload
        screenshot_path = None
        if 'screenshot_file' in request.files:
            file = request.files['screenshot_file']
            if file and file.filename and allowed_file(file.filename):
                try:
                    filename = secure_filename(file.filename)
                    # Add timestamp to prevent conflicts
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"{timestamp}_{filename}"
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(filepath)
                    screenshot_path = filename
                    logger.info(f"Image uploaded: {filename}")
                except Exception as e:
                    logger.error(f"File upload error: {str(e)}")
                    flash('Error uploading image. Please try again.', 'error')
        
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                
                cursor.execute('''
                    INSERT INTO bugs (title, description, steps, expected_result, actual_result, screenshot_url, screenshot_path, priority, created_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (title, description, steps, expected_result, actual_result, screenshot_url, screenshot_path, priority, session['user_id']))
                
                logger.info(f"New bug created by {session['user_email']}: {title}")
                flash('Bug reported successfully!', 'success')
                return redirect(url_for('dashboard'))
                
        except Exception as e:
            logger.error(f"Error creating bug: {str(e)}")
            flash('Error creating bug. Please try again.', 'error')
            return render_template('new_bug.html')
    
    return render_template('new_bug.html')

@app.route('/bug/<int:bug_id>')
@login_required
def view_bug(bug_id):
    """View bug details with comments"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Get bug details
            cursor.execute('''
                SELECT b.*, 
                       creator.email as creator_email,
                       assignee.email as assignee_email
                FROM bugs b
                LEFT JOIN users creator ON b.created_by = creator.id
                LEFT JOIN users assignee ON b.assigned_to = assignee.id
                WHERE b.id = ?
            ''', (bug_id,))
            bug = cursor.fetchone()
            
            if not bug:
                flash('Bug not found', 'error')
                return redirect(url_for('dashboard'))

            # Log view event
            try:
                log_bug_history(conn, bug_id, session.get('user_id'), 'viewed_bug', None, session.get('user_email'))
            except Exception as log_error:
                logger.warning(f"Failed to log view event for bug {bug_id}: {log_error}")
            
            # Get comments
            cursor.execute('''
                SELECT c.*, u.email as user_email
                FROM comments c
                JOIN users u ON c.user_id = u.id
                WHERE c.bug_id = ?
                ORDER BY c.created_at ASC
            ''', (bug_id,))
            comments = cursor.fetchall()
            
            # Get all users for assignment with roles
            cursor.execute('SELECT id, email, role FROM users ORDER BY email')
            users = cursor.fetchall()
            
            return render_template('view_bug.html', bug=bug, comments=comments, users=users)
        
    except Exception as e:
        logger.error(f"Error loading bug {bug_id}: {str(e)}")
        flash('Error loading bug details. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/bug/<int:bug_id>/comment', methods=['POST'])
@login_required
def add_comment(bug_id):
    """Add a comment to a bug"""
    comment_text = sanitize_input(request.form.get('comment', ''), 1000)
    
    if not comment_text or len(comment_text) < 2:
        flash('Comment must be at least 2 characters long', 'error')
        return redirect(url_for('view_bug', bug_id=bug_id))
    
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Verify bug exists
            cursor.execute('SELECT id FROM bugs WHERE id = ?', (bug_id,))
            if not cursor.fetchone():
                flash('Bug not found', 'error')
                return redirect(url_for('dashboard'))
            
            cursor.execute('''
                INSERT INTO comments (bug_id, user_id, comment)
                VALUES (?, ?, ?)
            ''', (bug_id, session['user_id'], comment_text))
            log_bug_history(conn, bug_id, session['user_id'], 'comment_added', None, comment_text)
            
            logger.info(f"Comment added to bug #{bug_id} by {session['user_email']}")
            flash('Comment added successfully!', 'success')
        
    except Exception as e:
        logger.error(f"Error adding comment: {str(e)}")
        flash('Error adding comment. Please try again.', 'error')
    
    return redirect(url_for('view_bug', bug_id=bug_id))

@app.route('/bug/<int:bug_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_bug(bug_id):
    """Edit an existing bug"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Get bug details
            cursor.execute('SELECT * FROM bugs WHERE id = ?', (bug_id,))
            bug = cursor.fetchone()
            
            if not bug:
                flash('Bug not found', 'error')
                return redirect(url_for('dashboard'))
            
            # Check permission (only creator or admin can edit)
            if bug['created_by'] != session['user_id'] and session.get('user_role') != 'admin':
                flash('You do not have permission to edit this bug', 'error')
                return redirect(url_for('view_bug', bug_id=bug_id))
            
            if request.method == 'POST':
                title = sanitize_input(request.form.get('title', ''), 200)
                description = sanitize_input(request.form.get('description', ''), 2000)
                steps = sanitize_input(request.form.get('steps', ''), 2000)
                screenshot_url = sanitize_input(request.form.get('screenshot_url', ''), 500)
                priority = request.form.get('priority', 'Medium')
                
                # Validation
                if not title or not description:
                    flash('Title and description are required', 'error')
                    return render_template('edit_bug.html', bug=bug)
                
                if priority not in ['Low', 'Medium', 'High']:
                    flash('Invalid priority level', 'error')
                    return render_template('edit_bug.html', bug=bug)
                
                # Handle file upload
                screenshot_path = bug['screenshot_path']
                if 'screenshot_file' in request.files:
                    file = request.files['screenshot_file']
                    if file and file.filename and allowed_file(file.filename):
                        try:
                            filename = secure_filename(file.filename)
                            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                            filename = f"{timestamp}_{filename}"
                            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                            file.save(filepath)
                            screenshot_path = filename
                        except Exception as e:
                            logger.error(f"File upload error: {str(e)}")
                
                # Log changes
                changes = []
                if bug['title'] != title:
                    changes.append(f"Title changed")
                    log_bug_history(conn, bug_id, session['user_id'], 'title_changed', bug['title'], title)
                if bug['description'] != description:
                    changes.append("Description updated")
                    log_bug_history(conn, bug_id, session['user_id'], 'description_changed')
                if bug['priority'] != priority:
                    changes.append(f"Priority: {bug['priority']} → {priority}")
                    log_bug_history(conn, bug_id, session['user_id'], 'priority_changed', bug['priority'], priority)
                
                # Update bug
                cursor.execute('''
                    UPDATE bugs 
                    SET title = ?, description = ?, steps = ?, screenshot_url = ?, 
                        screenshot_path = ?, priority = ?
                    WHERE id = ?
                ''', (title, description, steps, screenshot_url, screenshot_path, priority, bug_id))
                
                if changes:
                    log_bug_history(conn, bug_id, session['user_id'], 'bug_edited', ', '.join(changes))
                
                logger.info(f"Bug #{bug_id} updated by {session['user_email']}")
                flash('Bug updated successfully!', 'success')
                return redirect(url_for('view_bug', bug_id=bug_id))
            
            return render_template('edit_bug.html', bug=bug)
    
    except Exception as e:
        logger.error(f"Error editing bug: {str(e)}")
        flash('Error editing bug. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/bug/<int:bug_id>/assign', methods=['POST'])
@admin_required
def assign_bug(bug_id):
    """Assign a bug to a user (Admin only)"""
    assigned_to = request.form.get('assigned_to', '')
    
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Verify bug exists
            cursor.execute('SELECT id FROM bugs WHERE id = ?', (bug_id,))
            if not cursor.fetchone():
                flash('Bug not found', 'error')
                return redirect(url_for('dashboard'))
            
            if assigned_to:
                # Verify user exists
                cursor.execute('SELECT email FROM users WHERE id = ?', (assigned_to,))
                user = cursor.fetchone()
                if not user:
                    flash('Invalid user assignment', 'error')
                    return redirect(url_for('view_bug', bug_id=bug_id))
                
                cursor.execute('UPDATE bugs SET assigned_to = ? WHERE id = ?',
                             (assigned_to, bug_id))
                log_bug_history(conn, bug_id, session['user_id'], 'assigned_to', None, user['email'])
                logger.info(f"Bug #{bug_id} assigned to {user['email']} by {session['user_email']}")
            else:
                cursor.execute('UPDATE bugs SET assigned_to = NULL WHERE id = ?',
                             (bug_id,))
                log_bug_history(conn, bug_id, session['user_id'], 'assigned_to', None, 'Unassigned')
                logger.info(f"Bug #{bug_id} unassigned by {session['user_email']}")
            
            flash('Bug assignment updated!', 'success')
        
    except Exception as e:
        logger.error(f"Error assigning bug: {str(e)}")
        flash('Error updating assignment. Please try again.', 'error')
    
    return redirect(url_for('view_bug', bug_id=bug_id))

@app.route('/bug/<int:bug_id>/status', methods=['POST'])
@login_required
def update_status(bug_id):
    """Update bug status"""
    status = request.form.get('status', 'Open')
    status_note = sanitize_input(request.form.get('status_note', ''), 500)
    
    valid_statuses = ['Open', 'In Progress', 'Fixed', 'Closed']
    if status not in valid_statuses:
        flash('Invalid status', 'error')
        return redirect(url_for('view_bug', bug_id=bug_id))
    
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Verify bug exists and get current status
            cursor.execute('SELECT status FROM bugs WHERE id = ?', (bug_id,))
            bug = cursor.fetchone()
            if not bug:
                flash('Bug not found', 'error')
                return redirect(url_for('dashboard'))
            
            old_status = bug['status']
            cursor.execute('UPDATE bugs SET status = ? WHERE id = ?',
                         (status, bug_id))

            note_suffix = f" | note: {status_note}" if status_note else ""
            log_bug_history(conn, bug_id, session['user_id'], 'status_changed', old_status, f"{status}{note_suffix}")
            
            logger.info(f"Bug #{bug_id} status changed from '{old_status}' to '{status}' by {session['user_email']}")
            flash(f'Bug status updated to: {status}', 'success')
        
    except Exception as e:
        logger.error(f"Error updating status: {str(e)}")
        flash('Error updating status. Please try again.', 'error')
    
    return redirect(url_for('view_bug', bug_id=bug_id))

# ============== ADVANCED FEATURES ==============

@app.route('/api/bugs', methods=['GET'])
@login_required
def api_bugs():
    """API endpoint to get bugs in JSON format"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            status_filter = request.args.get('status', '')
            priority_filter = request.args.get('priority', '')
            
            query = '''
                SELECT b.id, b.title, b.description, b.priority, b.status, 
                       b.created_at, creator.email as creator_email
                FROM bugs b
                LEFT JOIN users creator ON b.created_by = creator.id
                WHERE 1=1
            '''
            params = []
            
            if status_filter:
                query += ' AND b.status = ?'
                params.append(status_filter)
            
            if priority_filter:
                query += ' AND b.priority = ?'
                params.append(priority_filter)
            
            query += ' ORDER BY b.created_at DESC'
            
            cursor.execute(query, params)
            bugs = [dict(row) for row in cursor.fetchall()]
            
            return jsonify({
                'success': True,
                'count': len(bugs),
                'bugs': bugs
            })
    
    except Exception as e:
        logger.error(f"API error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to fetch bugs'
        }), 500

@app.route('/bug/<int:bug_id>/delete', methods=['POST'])
@login_required
def delete_bug(bug_id):
    """Delete a bug (only if created by current user or admin)"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Check if user created this bug
            cursor.execute('SELECT created_by, title FROM bugs WHERE id = ?', (bug_id,))
            bug = cursor.fetchone()
            
            if not bug:
                flash('Bug not found', 'error')
                return redirect(url_for('dashboard'))
            
            if bug['created_by'] != session['user_id']:
                flash('You can only delete bugs you created', 'error')
                return redirect(url_for('view_bug', bug_id=bug_id))
            
            # Delete comments first (foreign key constraint)
            cursor.execute('DELETE FROM comments WHERE bug_id = ?', (bug_id,))
            
            # Delete bug
            cursor.execute('DELETE FROM bugs WHERE id = ?', (bug_id,))
            
            logger.info(f"Bug #{bug_id} '{bug['title']}' deleted by {session['user_email']}")
            flash('Bug deleted successfully', 'success')
            return redirect(url_for('dashboard'))
    
    except Exception as e:
        logger.error(f"Error deleting bug: {str(e)}")
        flash('Error deleting bug. Please try again.', 'error')
        return redirect(url_for('view_bug', bug_id=bug_id))

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    """Serve uploaded images"""
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/export/csv')
@login_required
def export_csv():
    """Export all bugs to CSV format with embedded images"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT b.id, b.title, b.description, b.priority, b.status,
                       b.created_at, b.screenshot_path, creator.email as creator,
                       assignee.email as assignee
                FROM bugs b
                LEFT JOIN users creator ON b.created_by = creator.id
                LEFT JOIN users assignee ON b.assigned_to = assignee.id
                ORDER BY b.created_at DESC
            ''')
            bugs = cursor.fetchall()
            
            # Create CSV content with image URLs
            csv_content = "ID,Title,Description,Priority,Status,Created At,Created By,Assigned To,Screenshot\n"
            for bug in bugs:
                screenshot_link = ""
                if bug['screenshot_path']:
                    # Create full URL for the image
                    screenshot_link = f"http://127.0.0.1:5000/uploads/{bug['screenshot_path']}"
                csv_content += f"{bug['id']},\"{bug['title']}\",\"{bug['description']}\",{bug['priority']},{bug['status']},{bug['created_at']},{bug['creator']},{bug['assignee'] or 'Unassigned'},{screenshot_link}\n"
            
            from flask import Response
            return Response(
                csv_content,
                mimetype="text/csv",
                headers={"Content-disposition": "attachment; filename=bugs_export.csv"}
            )
    
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        flash('Error exporting data. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/export/excel')
@login_required
def export_excel():
    """Export all bugs to Excel format with formatting"""
    try:
        # Try to import openpyxl, if not available, fall back to CSV
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            flash('Excel export not available. Please install openpyxl: pip install openpyxl', 'warning')
            return redirect(url_for('export_csv'))
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT b.id, b.title, b.description, b.priority, b.status,
                       b.created_at, b.screenshot_path, b.screenshot_url,
                       creator.email as creator, assignee.email as assignee
                FROM bugs b
                LEFT JOIN users creator ON b.created_by = creator.id
                LEFT JOIN users assignee ON b.assigned_to = assignee.id
                ORDER BY b.created_at DESC
            ''')
            bugs = cursor.fetchall()
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Bug Report"
            
            # Header style
            header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = ['ID', 'Title', 'Description', 'Priority', 'Status', 'Created At', 
                      'Created By', 'Assigned To', 'Screenshot URL']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data rows with formatting
            for row_num, bug in enumerate(bugs, 2):
                # Priority colors
                priority_colors = {
                    'High': 'fee2e2',
                    'Medium': 'fef3c7',
                    'Low': 'dcfce7'
                }
                
                # Status colors
                status_colors = {
                    'Open': 'fecaca',
                    'In Progress': 'fed7aa',
                    'Fixed': 'bbf7d0',
                    'Closed': 'd1d5db'
                }
                
                # Screenshot link
                screenshot_link = ""
                if bug['screenshot_path']:
                    screenshot_link = f"http://127.0.0.1:5000/uploads/{bug['screenshot_path']}"
                elif bug['screenshot_url']:
                    screenshot_link = bug['screenshot_url']
                
                data = [
                    bug['id'],
                    bug['title'],
                    bug['description'],
                    bug['priority'],
                    bug['status'],
                    bug['created_at'],
                    bug['creator'],
                    bug['assignee'] or 'Unassigned',
                    screenshot_link
                ]
                
                for col_num, value in enumerate(data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    # Apply priority color
                    if col_num == 4:  # Priority column
                        cell.fill = PatternFill(start_color=priority_colors.get(value, 'FFFFFF'),
                                               end_color=priority_colors.get(value, 'FFFFFF'),
                                               fill_type="solid")
                        cell.font = Font(bold=True)
                    
                    # Apply status color
                    if col_num == 5:  # Status column
                        cell.fill = PatternFill(start_color=status_colors.get(value, 'FFFFFF'),
                                               end_color=status_colors.get(value, 'FFFFFF'),
                                               fill_type="solid")
                        cell.font = Font(bold=True)
                    
                    # Make screenshot URL clickable
                    if col_num == 9 and value:
                        cell.hyperlink = value
                        cell.font = Font(color="0563C1", underline="single")
            
            # Adjust column widths
            column_widths = [8, 30, 50, 12, 15, 20, 25, 25, 40]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col_num)].width = width
            
            # Row heights
            ws.row_dimensions[1].height = 25
            for row in range(2, len(bugs) + 2):
                ws.row_dimensions[row].height = 60
            
            # Save to BytesIO
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            from flask import send_file
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'bugs_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
    
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        flash('Error exporting to Excel. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/api/check-duplicates', methods=['POST'])
@login_required
def check_duplicates():
    """Check for similar bug titles to detect potential duplicates"""
    try:
        data = request.get_json()
        title = sanitize_input(data.get('title', ''), 200).lower()
        
        if len(title) < 5:
            return jsonify({'similar_bugs': []})
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Simple similarity check - find bugs with matching words
            words = set(title.split())
            if len(words) < 2:
                return jsonify({'similar_bugs': []})
            
            cursor.execute('''
                SELECT id, title, status, priority, created_at 
                FROM bugs 
                WHERE status != 'Closed'
                ORDER BY created_at DESC 
                LIMIT 100
            ''')
            all_bugs = cursor.fetchall()
            
            similar_bugs = []
            for bug in all_bugs:
                bug_words = set(bug['title'].lower().split())
                # Calculate word overlap
                overlap = len(words.intersection(bug_words))
                if overlap >= min(2, len(words) * 0.5):  # At least 2 words or 50% match
                    similar_bugs.append({
                        'id': bug['id'],
                        'title': bug['title'],
                        'status': bug['status'],
                        'priority': bug['priority'],
                        'created_at': bug['created_at'][:16]
                    })
                    if len(similar_bugs) >= 5:
                        break
            
            return jsonify({'similar_bugs': similar_bugs})
    
    except Exception as e:
        logger.error(f"Duplicate check error: {str(e)}")
        return jsonify({'similar_bugs': []})

@app.route('/bug/<int:bug_id>/history')
@login_required
def bug_history(bug_id):
    """View complete history of a bug"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Get bug details
            cursor.execute('SELECT * FROM bugs WHERE id = ?', (bug_id,))
            bug = cursor.fetchone()
            
            if not bug:
                flash('Bug not found', 'error')
                return redirect(url_for('dashboard'))
            
            # Get history records
            cursor.execute('''
                SELECT bh.*, u.email as user_email
                FROM bug_history bh
                LEFT JOIN users u ON bh.user_id = u.id
                WHERE bh.bug_id = ?
                ORDER BY bh.created_at DESC
            ''', (bug_id,))
            history = cursor.fetchall()
            
            return render_template('bug_history.html', bug=bug, history=history)
    
    except Exception as e:
        logger.error(f"History view error: {str(e)}")
        flash('Error loading history. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/health')
def health_check():
    """Health check endpoint for monitoring"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) as count FROM bugs')
            cursor.fetchone()
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        logger.error(f"Health check failed: {str(e)}")
        return jsonify({
            'status': 'unhealthy',
            'error': str(e)
        }), 500

# ============== APPLICATION STARTUP ==============

if __name__ == '__main__':
    # Initialize database on first run
    if not os.path.exists(DATABASE):
        print("[SETUP] First time setup - Creating database...")
        init_db()
    else:
        print("[OK] Database found - connecting...")
        # Ensure all tables exist even if database exists
        init_db()
    
    print("[START] Starting Bug Reporting Tool v3.1.0...")
    print("[INFO] Access the application at: http://127.0.0.1:5000")
    print("=" * 50)
    
    app.run(debug=True, host='0.0.0.0', port=5000)

